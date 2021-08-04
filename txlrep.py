# -*- coding: utf-8 -*-
"""Модуль роутинга главной страницы. """  # noqa

# from openpyxl import Workbook
from openpyxl import styles as xlstyles

THIN_BORDER_STYLE = xlstyles.Side(border_style="thin", color="000000")
TOP_BORDER = xlstyles.Border(top=THIN_BORDER_STYLE)
BOTTOM_BORDER = xlstyles.Border(bottom=THIN_BORDER_STYLE)
LEFT_BORDER = xlstyles.Border(left=THIN_BORDER_STYLE)
RIGHT_BORDER = xlstyles.Border(right=THIN_BORDER_STYLE)

COLUMN_A = 1
COLUMN_B = 2
COLUMN_C = 3
COLUMN_D = 4
COLUMN_E = 5
COLUMN_F = 6
COLUMN_G = 7
COLUMN_H = 8
COLUMN_I = 9
COLUMN_J = 10
COLUMN_K = 11
COLUMN_L = 12
COLUMN_M = 13
COLUMN_N = 14
COLUMN_O = 15
COLUMN_P = 16
COLUMN_Q = 17
COLUMN_R = 18
COLUMN_S = 19
COLUMN_T = 20
COLUMN_U = 21
COLUMN_V = 22
COLUMN_W = 23
COLUMN_X = 24
COLUMN_Y = 25
COLUMN_Z = 26


def write_in_cell(psheet, pcell, pvalue, pfont, palign):
    """Выводит в заданную ячейку значение."""
    assert psheet is not None, ("Assert: [txlrep:write_in_cell]:"
                                "No <psheet> parameter specified!")
    assert pcell is not None, ("Assert: [txlrep:write_in_cell]:"
                               "No <pcell> parameter specified!")
    assert pvalue is not None, ("Assert: [txlrep:write_in_cell]:"
                                "No <pwhat> parameter specified!")
    assert pfont is not None, ("Assert: [txlrep:write_in_cell]:"
                               "No <pfont> parameter specified!")
    assert palign is not None, ("Assert: [txlrep:write_in_cell]:"
                                "No <palign> parameter specified!")
    cell = psheet[pcell]
    cell.alignment = palign
    cell.font = pfont
    cell.value = pvalue
    return cell


# py lint: disable=too-many-statements
# py lint: disable=too-many-locals
# py lint: disable=too-many-arguments
def write_in_cell_by_row_col(psheet, prow, pcolumn, pvalue, pfont, palign):
    """Выводит в заданную ячейку значение."""
    assert psheet is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                                "No <psheet> parameter specified!")
    assert prow is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                              "No <prow> parameter specified!")
    assert pcolumn is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                                 "No <pcolumn> parameter specified!")
    assert pvalue is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                                "No <pvalue> parameter specified!")
    assert pfont is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                               "No <pfont> parameter specified!")
    assert palign is not None, ("Assert: [txlrep:write_in_cell_by_row_col]:"
                                "No <palign> parameter specified!")
    cell = psheet.cell(column=pcolumn, row=prow)
    cell.alignment = palign
    cell.font = pfont
    cell.value = pvalue
    return cell


def merge_and_write(psheet, pfrom, pto, pvalue, pfont, palign):
    """Объединяет ячейки и выводит информацию в них."""
    assert psheet is not None, ("Assert: [txlrep:merge_and_write]:"
                                "No <psheet> parameter specified!")
    assert pfrom is not None, ("Assert: [txlrep:merge_and_write]:"
                               "No <pfrom> parameter specified!")
    assert pto is not None, ("Assert: [txlrep:merge_and_write]:"
                             "No <pto> parameter specified!")
    assert pvalue is not None, ("Assert: [txlrep:merge_and_write]:"
                                "No <pvalue> parameter specified!")
    assert pfont is not None, ("Assert: [txlrep:merge_and_write]:"
                               "No <pfont> parameter specified!")
    assert palign is not None, ("Assert: [txlrep:merge_and_write]:"
                                "No <palign> parameter specified!")
    psheet.merge_cells(pfrom + ":" + pto)
    cell = psheet[pfrom]
    cell.alignment = palign
    cell.font = pfont
    cell.value = pvalue
    return cell


def draw_horizontal_line(psheet, prow, pcolumn_begin, pcolumn_end,
                         ptop_line=True):
    """Рисует горизонтальную границу."""
    assert psheet is not None, ("Assert: [txlrep:draw_horizontal_line]:"
                                "No <psheet> parameter specified!")
    assert prow is not None, ("Assert: [txlrep:draw_horizontal_line]:"
                              "No <prow> parameter specified!")
    assert pcolumn_begin is not None, ("Assert: [txlrep:"
                                       "draw_horizontal_line]:"
                                       "No <pcolumn_begin> parameter"
                                       " specified!")
    assert pcolumn_end is not None, ("Assert: [txlrep:draw_horizontal_line]:"
                                     "No <pcolumn_end> parameter specified!")
    assert ptop_line is not None, ("Assert: [txlrep:draw_horizontal_line]:"
                                   "No <ptop_line> parameter specified!")
    for column in range(pcolumn_begin, pcolumn_end + 1):

        cell = psheet.cell(column=column, row=prow)
        if ptop_line:

            cell.border = cell.border + TOP_BORDER
        else:

            cell.border = cell.border + BOTTOM_BORDER


def draw_vertical_line(psheet, pcolumn, prow_begin, prow_end, pleft_line=True):
    """Рисует вертикальную границу."""
    assert psheet is not None, ("Assert: [txlrep:draw_vertical_line]:"
                                "No <psheet> parameter specified!")
    assert pcolumn is not None, ("Assert: [txlrep:draw_vertical_line]:"
                                 "No <pcolumn> parameter specified!")
    assert prow_begin is not None, ("Assert: [txlrep:"
                                    "draw_vertical_line]:"
                                    "No <prow_begin> parameter"
                                    " specified!")
    assert prow_end is not None, ("Assert: [txlrep:draw_vertical_line]:"
                                  "No <prow_end> parameter specified!")
    assert pleft_line is not None, ("Assert: [txlrep:draw_vertical_line]:"
                                    "No <pleft_line> parameter specified!")
    for row in range(prow_begin, prow_end + 1):

        cell = psheet.cell(column=pcolumn, row=row)
        if pleft_line:

            cell.border = cell.border + LEFT_BORDER
        else:

            cell.border = cell.border + RIGHT_BORDER


def frame_cell(pcell, ptop=None, pleft=None, pbottom=None,
               pright=None):
    """Рисует прямоугольник вокруг ячейки."""
    assert pcell is not None, ("Assert: [txlrep:frame_cell]:"
                               "No <pcell> parameter specified!")
    """Задает границы ячейки."""
    if ptop:

        pcell.border = pcell.border + ptop
    if pleft:

        pcell.border = pcell.border + pleft
    if pbottom:

        pcell.border = pcell.border + pbottom
    if pright:

        pcell.border = pcell.border + pright
    return pcell
